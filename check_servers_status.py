# -*- coding: utf-8 -*-
import openpyxl as openpyxl
import xlrd
import os

workbook = openpyxl.Workbook()
sheet = workbook.active
row_index = 1
next_server_row_index = 1

workbook1 = openpyxl.Workbook()
sheet1 = workbook1.active
row_index1 = 1
next_server_row_index1 = 1



def get_iftop(serverName):
    global sheet1
    global row_index1
    global next_server_row_index1
    iftop_row_index = row_index1

    sheet1.cell(row=iftop_row_index, column=1, value=serverName)

    command = "ssh %s \'sudo iftop -i eth0 -t -s 1\'"%serverName

    try:
        p = os.popen(command)
        lines = p.readlines()
        # print len(lines)
        line = lines[1].split()
        if line[11] == "40s":
            data_type = 0
            for v in range(3,len(lines)-1):
                if lines[v].startswith('--'):
                    break
                if lines[v].startswith('--'):
                    data_type += 1
                    continue
                if data_type == 0:
                    line_now = lines[v].split()
                    # print line_now
                    if str.isdigit(line_now[0]):
                        sheet1.cell(row=iftop_row_index, column=2, value=line_now[0])
                        sheet1.cell(row=iftop_row_index, column=3, value=line_now[1])
                        sheet1.cell(row=iftop_row_index, column=4, value="\\"+line_now[2])
                        sheet1.cell(row=iftop_row_index, column=5, value=line_now[5])
                        iftop_row_index += 1
                    else:
                        sheet1.cell(row=iftop_row_index, column=3, value=line_now[0])
                        sheet1.cell(row=iftop_row_index, column=4, value=line_now[1])
                        sheet1.cell(row=iftop_row_index, column=5, value=line_now[4])
                        iftop_row_index += 1
                elif data_type > 0:
                    line_now = filter(None, lines[v].split("  "))
                    # print line_now
                    sheet1.cell(row=iftop_row_index, column=2, value=line_now[0])
                    sheet1.cell(row=iftop_row_index, column=5, value=line_now[1])
                    iftop_row_index += 1
    except Exception as e:
        iftop_row_index += 1
        print e
    next_server_row_index1 = iftop_row_index
    print "next_server_row_index1:",next_server_row_index1


def get_iostat(serverName):
    global sheet
    global row_index
    global next_server_row_index
    iostat_row_index = row_index

    sheet.cell(row=iostat_row_index, column=1, value=serverName)

    command = "ssh %s \'iostat\'"%serverName
    p = os.popen(command)
    lines = p.readlines()
    # print len(lines)
    try:
        for v in range(0,len(lines)):
            if v == 0:
                cores = lines[v].split("(")
                cores1 = cores[len(cores)-1].split(')')[0]
                sheet.cell(row=iostat_row_index, column=5, value=cores1)
            elif v == 3:
                avg_cpu = lines[v].split()
                sheet.cell(row=iostat_row_index, column=3, value=avg_cpu[3])
                sheet.cell(row=iostat_row_index, column=4, value=avg_cpu[5])

        iostat_row_index += 1
    except Exception as e:
        print e
    if iostat_row_index > next_server_row_index:
        next_server_row_index = iostat_row_index
    print "next_server_row_index:",next_server_row_index


def get_free_mem(serverName):
    global sheet
    global row_index
    global next_server_row_index
    free_row_index = row_index

    command = "ssh %s \'free -m\'"%serverName
    p = os.popen(command)
    lines = p.readlines()
    try:
        line = lines[1].split()
        for v in range(1,len(line)):
            sheet.cell(row=free_row_index, column=6 + v, value=line[v])
        free_row_index += 1
    except Exception as e:
        print e
    if free_row_index > next_server_row_index:
        next_server_row_index = free_row_index
    print "next_server_row_index:",next_server_row_index


if __name__ == '__main__':

    # sheet1.merge_cells('B1:E1')
    sheet1.cell(row=row_index1, column=2, value="iftop")
    row_index1 += 1
    sheet1.cell(row=row_index1, column=3, value="Host name (port/service if enabled)")
    sheet1.cell(row=row_index1, column=5, value="last 40s")
    row_index1 += 1

    # sheet.merge_cells('B1:E1')
    sheet.cell(row=row_index, column=2, value="iostat")
    # sheet.merge_cells('F1:K1')
    sheet.cell(row=row_index, column=6, value="free -m")
    row_index += 1
    sheet.cell(row=row_index, column=2, value="avg-cpu:")
    sheet.cell(row=row_index, column=3, value="%iowait")
    sheet.cell(row=row_index, column=4, value="%idle")
    sheet.cell(row=row_index, column=5, value="cores")
    sheet.cell(row=row_index, column=6, value="Mem:")
    sheet.cell(row=row_index, column=7, value="total")
    sheet.cell(row=row_index, column=8, value="used")
    sheet.cell(row=row_index, column=9, value="free")
    sheet.cell(row=row_index, column=10, value="shared")
    sheet.cell(row=row_index, column=11, value="buffers")
    sheet.cell(row=row_index, column=12, value="cached")
    row_index += 1
    # 打开excel文件
    data1 = xlrd.open_workbook("./ecs_instance_list_cn-beijing_2019-12-03.xlsx")
    # 读取第一个工作表
    table = data1.sheets()[0]
    # 统计行数
    n_rows = table.nrows
    print n_rows

    for v in range(2, n_rows):
        values = table.row_values(v)[1]
        print values
        # values = "abj-gateway-1"
        get_iftop(values)
        get_iostat(values)
        get_free_mem(values)
        row_index = next_server_row_index
        row_index1 = next_server_row_index1

    workbook.save("./iostat_and_mem_data.xlsx")
    workbook1.save("./iftop_data.xlsx")


# def write_excel_xlsx(path, sheet_name, value):
#     index = len(value)
#     workbook = openpyxl.Workbook()
#     sheet = workbook.active
#     sheet.title = sheet_name
#     for i in range(0, index):
#         for j in range(0, len(value[i])):
#             sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
#     workbook.save(path)
#     print("xlsx格式表格写入数据成功！")
#
